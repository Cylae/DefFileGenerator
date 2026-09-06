#!/usr/bin/env python3
"""
Modbus Register Extractor Module.

Provides format-agnostic extraction of register tables from Excel, PDF, CSV, and XML files,
applying two-pass heuristic column mapping, address offset adjustments,
and lazy generator streaming.
"""

import argparse
import csv
import itertools
import json
import logging
import os
import sys
import zipfile
from collections.abc import Iterable, Iterator
from typing import Any, Optional, Union

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PDF_ERRORS: tuple[type[Exception], ...]
try:
    import pdfplumber

    HAS_PDFPLUMBER = True
    try:
        from pdfminer.pdfparser import PDFSyntaxError
        from pdfplumber.utils.exceptions import PdfminerException

        PDF_ERRORS = (PDFSyntaxError, PdfminerException)
    except ImportError:
        PDF_ERRORS = ()
except ImportError:
    HAS_PDFPLUMBER = False
    PDF_ERRORS = ()

SECURITY_EXCEPTIONS: tuple[type[Exception], ...]
try:
    from defusedxml import ElementTree as ET
    from defusedxml.common import DTDForbidden, EntitiesForbidden, ExternalReferenceForbidden

    HAS_DEFUSEDXML = True
    SECURITY_EXCEPTIONS = (EntitiesForbidden, DTDForbidden, ExternalReferenceForbidden)
except ImportError:
    HAS_DEFUSEDXML = False
    SECURITY_EXCEPTIONS = ()

XML_PARSE_ERRORS: tuple[type[Exception], ...]
try:
    import xml.etree.ElementTree as ET_STD

    XML_PARSE_ERRORS = (ET_STD.ParseError,)
except ImportError:
    XML_PARSE_ERRORS = ()

# Secure Import for Generator and peek_generator with clean fallbacks
peek_generator: Any
Generator: Any
try:
    from DefFileGenerator.def_gen import Generator, peek_generator
except ImportError:
    try:
        from def_gen import Generator, peek_generator  # type: ignore[import-not-found, no-redef]
    except ImportError:
        Generator = None
        peek_generator = None

if peek_generator is None:

    def peek_generator(iterable: Optional[Iterable]) -> tuple[bool, Iterator]:
        """Checks if an iterable is non-empty without fully consuming it.

        Returns (has_data, original_iterator).
        """
        if iterable is None:
            return False, iter([])
        it = iter(iterable)
        try:
            first = next(it)
        except StopIteration:
            return False, iter([])
        return True, itertools.chain([first], it)


class Extractor:
    COLUMN_MAPPING: dict[str, list[str]] = {
        "RegisterType": ["register type", "reg type", "modbus type", "registertype"],
        "Address": ["address", "addr", "offset", "register", "reg"],
        "Name": ["name", "description", "parameter", "variable", "signal", "signal name"],
        "Type": ["data type", "datatype", "type", "format"],
        "Unit": ["unit", "units"],
        "Tag": ["tag"],
        "Action": ["action", "access"],
        "Factor": ["scale", "factor", "multiplier", "ratio"],
        "Offset": ["offset", "bias", "coefficient b"],
        "ScaleFactor": ["scalefactor", "scale factor"],
        "Length": ["length", "len", "size", "count", "quantity"],
        "StartBit": ["startbit", "bit offset", "bit", "start"],
    }

    # Detection is ordered most-specific-first so that greedy substring matches
    # (e.g. 'offset' for both Address and Offset) resolve deterministically.
    DETECTION_ORDER = (
        "RegisterType",
        "Address",
        "Name",
        "Type",
        "Unit",
        "Action",
        "Tag",
        "Factor",
        "Offset",
        "ScaleFactor",
        "Length",
        "StartBit",
    )

    def __init__(self, mapping: Optional[dict[str, str]] = None) -> None:
        self.mapping = mapping or {}

    @staticmethod
    def normalize_type(t: Any) -> str:
        if Generator is not None:
            return Generator.normalize_type(t)
        return str(t).upper() if t else "U16"

    def extract_from_excel(
        self, filepath: str, sheet_name: Optional[str] = None
    ) -> Iterator[Iterator[dict[str, Any]]]:
        if not HAS_OPENPYXL:
            logging.error("openpyxl is required for Excel extraction.")
            return iter([])

        def excel_sheets_generator() -> Iterator[Iterator[dict[str, Any]]]:
            if sheet_name:
                def sheet_generator() -> Iterator[dict[str, Any]]:
                    wb = None
                    try:
                        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                        ws = wb[sheet_name]
                        rows = ws.iter_rows(values_only=True)
                        try:
                            header_row = next(rows)
                        except StopIteration:
                            return
                        headers = [str(h).strip() if h is not None else "" for h in header_row]
                        for row in rows:
                            if any(cell is not None and str(cell).strip() for cell in row):
                                yield {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
                    except (OSError, zipfile.BadZipFile) as e:
                        logging.error(f"File IO Error extracting from Excel {filepath}: {e}")
                    except (ValueError, TypeError, KeyError) as e:
                        logging.error(f"Error extracting from Excel {filepath}: {e}")
                    finally:
                        if wb is not None:
                            wb.close()

                yield sheet_generator()
            else:
                sheet_names = None
                wb_outer = None
                try:
                    wb_outer = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                    sheet_names = wb_outer.sheetnames
                except (OSError, zipfile.BadZipFile, ValueError, TypeError, KeyError):
                    pass
                finally:
                    if wb_outer is not None:
                        wb_outer.close()

                if sheet_names is None:
                    def error_sheet_generator() -> Iterator[dict[str, Any]]:
                        try:
                            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                            wb.close()
                        except (OSError, zipfile.BadZipFile) as e:
                            logging.error(f"File IO Error extracting from Excel {filepath}: {e}")
                        except (ValueError, TypeError, KeyError) as e:
                            logging.error(f"Error extracting from Excel {filepath}: {e}")
                        return iter([])
                        yield

                    yield error_sheet_generator()
                    return

                for sname in sheet_names:
                    def make_sheet_gen(s_name=sname):
                        def sheet_generator() -> Iterator[dict[str, Any]]:
                            wb_inner = None
                            try:
                                wb_inner = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                                ws = wb_inner[s_name]
                                rows = ws.iter_rows(values_only=True)
                                try:
                                    header_row = next(rows)
                                except StopIteration:
                                    return
                                headers = [str(h).strip() if h is not None else "" for h in header_row]
                                for row in rows:
                                    if any(cell is not None and str(cell).strip() for cell in row):
                                        yield {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
                            except (OSError, zipfile.BadZipFile) as e:
                                logging.error(f"File IO Error extracting from Excel {filepath}: {e}")
                            except (ValueError, TypeError, KeyError) as e:
                                logging.error(f"Error extracting from Excel {filepath}: {e}")
                            finally:
                                if wb_inner is not None:
                                    wb_inner.close()
                        return sheet_generator()

                    yield make_sheet_gen()

        return excel_sheets_generator()

    def extract_from_pdf(
        self, filepath: str, pages: Optional[Union[int, list[Union[int, str]], str]] = None
    ) -> Iterator[Iterator[dict[str, Any]]]:
        if not HAS_PDFPLUMBER:
            logging.error("pdfplumber is required for PDF extraction.")
            return iter([])

        def pdf_tables_generator() -> Iterator[Iterator[dict[str, Any]]]:
            try:
                with pdfplumber.open(filepath) as pdf:
                    target_pages = []
                    if pages is None:
                        target_pages = pdf.pages
                    else:
                        requested = pages if isinstance(pages, list) else [pages]
                        if isinstance(pages, str):
                            requested = [p.strip() for p in pages.split(",")]
                        for p in requested:
                            try:
                                idx = int(p) - 1
                                if 0 <= idx < len(pdf.pages):
                                    target_pages.append(pdf.pages[idx])
                                else:
                                    logging.warning(
                                        f"Page {p} is out of range (1-{len(pdf.pages)})"
                                    )
                            except (ValueError, TypeError):
                                logging.warning(f"Invalid page reference: {p}")

                    for page in target_pages:
                        tables = page.extract_tables()
                        for table in tables:
                            if not table or len(table) < 2:
                                continue

                            def table_generator(current_table=table) -> Iterator[dict[str, Any]]:
                                headers = [
                                    str(c).replace("\n", " ").strip() if c else ""
                                    for c in current_table[0]
                                ]
                                for row in current_table[1:]:
                                    row_dict = {}
                                    for i, cell in enumerate(row):
                                        if i < len(headers):
                                            row_dict[headers[i]] = (
                                                str(cell).replace("\n", " ").strip() if cell else ""
                                            )
                                    if any(v.strip() for v in row_dict.values() if v):
                                        yield row_dict

                            yield table_generator()
            except (OSError,) + PDF_ERRORS as e:
                logging.error(
                    f"File IO Error or PDF Syntax Error extracting from PDF {filepath}: {e}"
                )
            except (ValueError, TypeError, IndexError) as e:
                logging.error(f"Error extracting from PDF {filepath}: {e}")

        return pdf_tables_generator()

    def extract_from_csv(self, filepath: str) -> Iterator[Iterator[dict[str, Any]]]:
        def csv_tables_generator() -> Iterator[Iterator[dict[str, Any]]]:
            def csv_table_generator() -> Iterator[dict[str, Any]]:
                try:
                    with open(filepath, "rb") as f:
                        header_bytes = f.read(4)
                        encoding = (
                            "utf-16"
                            if header_bytes.startswith((b"\xff\xfe", b"\xfe\xff"))
                            else "utf-8-sig"
                        )

                    with open(filepath, encoding=encoding) as f:
                        snippet = f.read(2048)
                        f.seek(0)
                        try:
                            dialect = csv.Sniffer().sniff(snippet, delimiters=";,")
                            delimiter = dialect.delimiter
                        except csv.Error:
                            delimiter = ","
                            for d in [",", ";", "\t"]:
                                if d in snippet:
                                    delimiter = d
                                    break

                        reader = csv.DictReader(f, delimiter=delimiter)
                        for row in reader:
                            if any(val.strip() for val in row.values() if val is not None):
                                yield dict(row)
                except OSError as e:
                    logging.error(f"File IO Error extracting from CSV {filepath}: {e}")
                except csv.Error as e:
                    logging.error(f"CSV Parsing Error in {filepath}: {e}")
                except UnicodeError as e:
                    logging.error(f"Encoding Error extracting from CSV {filepath}: {e}")
                except (ValueError, TypeError) as e:
                    logging.error(f"Unexpected error extracting from CSV {filepath}: {e}")

            yield csv_table_generator()

        return csv_tables_generator()

    def extract_from_xml(self, filepath: str) -> Iterator[Iterator[dict[str, Any]]]:
        if not HAS_DEFUSEDXML:
            logging.error("defusedxml is required for secure XML parsing.")
            return iter([])

        def xml_tables_generator() -> Iterator[Iterator[dict[str, Any]]]:
            def xml_generator() -> Iterator[dict[str, Any]]:
                try:
                    f = open(filepath, "rb")
                    try:
                        tree = ET.parse(f)
                        root = tree.getroot()
                    finally:
                        f.close()
                    seen = set()
                    for elem in root.iter():
                        row = {}
                        for child in elem:
                            if len(child) == 0 and child.text:
                                row[child.tag] = child.text.strip()
                        if len(row) >= 2:
                            js = json.dumps(row, sort_keys=True)
                            if js not in seen:
                                seen.add(js)
                                yield row
                except SECURITY_EXCEPTIONS as e:
                    logging.error(f"Security error parsing XML {filepath}: {e}")
                    raise
                except (OSError,) + XML_PARSE_ERRORS as e:
                    logging.error(
                        f"File IO Error or Parsing Error extracting from XML {filepath}: {e}"
                    )
                except (ValueError, TypeError) as e:
                    logging.error(f"Error extracting from XML {filepath}: {e}")

            # Return as a single table
            yield xml_generator()

        return xml_tables_generator()

    def map_and_clean(
        self, tables: Optional[Iterable[Iterable[dict[str, Any]]]], address_offset: int = 0
    ) -> Iterator[dict[str, Any]]:
        if not tables:
            return

        for table in tables:
            has_rows, table = peek_generator(table)
            if not has_rows:
                continue

            iterator = iter(table)
            buffer = []
            try:
                for _ in range(50):
                    buffer.append(next(iterator))
            except StopIteration:
                pass

            if not buffer:
                continue

            # Insertion-ordered, de-duplicated column list. A plain set was
            # used here previously, and because str hashing is randomised per
            # interpreter (PYTHONHASHSEED), detection could bind a different
            # source column between runs whenever two targets share a pattern
            # (e.g. 'offset' matches both Address and Offset). That made the
            # generated definition non-reproducible; ordering fixes it.
            all_keys = list(dict.fromkeys(key for row in buffer for key in row.keys()))

            col_map = {}
            used_src_cols = set()

            for target, source in self.mapping.items():
                if source in all_keys:
                    col_map[target] = source
                    used_src_cols.add(source)

            # Lower-case each source column exactly once instead of once per
            # (target, pattern) probe: this is O(K) rather than O(K * T * P).
            lowered = {src: str(src).lower().strip() for src in all_keys}

            # Pass 1: exact header match. Pass 2: substring heuristic.
            for exact_pass in (True, False):
                for target in self.DETECTION_ORDER:
                    if target in col_map:
                        continue
                    patterns = self.COLUMN_MAPPING.get(target, (target.lower(),))
                    for src_col in all_keys:
                        if src_col in used_src_cols:
                            continue
                        s_low = lowered[src_col]
                        matched = (
                            s_low in patterns
                            if exact_pass
                            else any(pat in s_low for pat in patterns)
                        )
                        if matched:
                            col_map[target] = src_col
                            used_src_cols.add(src_col)
                            break

            def process_row(
                r: dict[str, Any], col_map: dict[str, Any] = col_map
            ) -> Optional[dict[str, Any]]:
                new_row = {target: r.get(src_col) for target, src_col in col_map.items()}
                if not new_row.get("Name") and not new_row.get("Address"):
                    return None

                sbit_col = col_map.get("StartBit")
                slen_col = col_map.get("Length")
                sbit_val = r.get(sbit_col) if sbit_col else None
                slen_val = r.get(slen_col) if slen_col else None
                sbit = str(sbit_val).strip() if sbit_val is not None else ""
                slen = str(slen_val).strip() if slen_val is not None else ""

                dtype = self.normalize_type(new_row.get("Type", "U16"))
                new_row["Type"] = dtype

                addr = str(new_row.get("Address", "")).strip()
                if dtype == "BITS" and sbit != "" and "_" not in addr:
                    if slen == "":
                        slen = "1"
                    addr = f"{addr}_{sbit}_{slen}"
                elif (
                    (dtype == "STRING" or dtype.startswith("STR"))
                    and slen != ""
                    and "_" not in addr
                ):
                    addr = f"{addr}_{slen}"
                if Generator is not None:
                    new_row["Address"] = Generator.apply_address_offset(addr, address_offset)
                else:
                    new_row["Address"] = addr

                if new_row.get("Factor") is not None:
                    if Generator is not None:
                        new_row["Factor"] = str(Generator._parse_numeric(new_row["Factor"], 1.0))
                if "RegisterType" not in new_row or not new_row["RegisterType"]:
                    new_row["RegisterType"] = "Holding Register"
                return new_row

            for row in itertools.chain(buffer, iterator):
                processed = process_row(row)
                if processed is not None:
                    yield processed


def main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    parser = argparse.ArgumentParser(description="Extract register information.")
    parser.add_argument("input_file")
    parser.add_argument("-o", "--output")
    parser.add_argument("--mapping")
    parser.add_argument("--sheet")
    parser.add_argument("--pages")
    parser.add_argument("--address-offset", type=int, default=0)
    args = parser.parse_args()
    mapping = {}
    if args.mapping:
        with open(args.mapping) as f:
            mapping = json.load(f)
    extractor = Extractor(mapping)
    ext = os.path.splitext(args.input_file)[1].lower()
    pages = args.pages
    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        raw = extractor.extract_from_excel(args.input_file, args.sheet)
    elif ext == ".pdf":
        raw = extractor.extract_from_pdf(args.input_file, pages)
    elif ext == ".csv":
        raw = extractor.extract_from_csv(args.input_file)
    elif ext == ".xml":
        raw = extractor.extract_from_xml(args.input_file)
    else:
        logging.error(f"Unsupported extension: {ext}")
        sys.exit(1)
    mapped = list(extractor.map_and_clean(raw, args.address_offset))
    out = open(args.output, "w", newline="", encoding="utf-8") if args.output else sys.stdout
    writer = csv.DictWriter(
        out,
        fieldnames=[
            "Name",
            "Tag",
            "RegisterType",
            "Address",
            "Type",
            "Factor",
            "Offset",
            "Unit",
            "Action",
            "ScaleFactor",
        ],
        extrasaction="ignore",
    )
    writer.writeheader()
    writer.writerows(mapped)
    if args.output:
        out.close()


if __name__ == "__main__":
    main()
